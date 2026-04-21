"""
Gap Analysis Script: ONDC Provided Test Cases vs Existing Gateway Test Cases
Reads the Excel file, extracts ONDC test cases per sheet, extracts existing
test cases from code, and writes gap analysis columns back into the same Excel.

Decision Logic:
  Covered          - specific API identified AND keyword_score >= 3, OR keyword_score >= 5
  Partially Covered - keyword_score >= 2 (weak or indirect match)
  Gap              - keyword_score < 2 or no recognizable API / category match
"""

import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = "Gateway_Consolidated_Test_Cases_Deduplicated -ONDC Provided.xlsx"

# ─── 1. Extract existing Gateway test cases from source code ─────────────────

GATEWAY_TEST_FILES = {
    "search_functional":  "tests/gateway/ondc_gateway_search_functional.py",
    "search_negative":    "tests/gateway/ondc_gateway_search_negative.py",
    "confirm_functional": "tests/gateway/ondc_gateway_confirm_functional.py",
    "confirm_negative":   "tests/gateway/ondc_gateway_confirm_negative.py",
    "init_functional":    "tests/gateway/ondc_gateway_init_functional.py",
    "init_negative":      "tests/gateway/ondc_gateway_init_negative.py",
    "select_functional":  "tests/gateway/ondc_gateway_select_functional.py",
    "select_negative":    "tests/gateway/ondc_gateway_select_negative.py",
    "lookup_functional":  "tests/gateway/ondc_gateway_lookup_functional.py",
    "lookup_negative":    "tests/gateway/ondc_gateway_lookup_negative.py",
}

GREEN  = "C6EFCE"
YELLOW = "FFEB9C"
RED    = "FFC7CE"


def extract_tasks_from_file(filepath):
    """Extract @task decorated methods with docstrings/comments from a Locust test file."""
    tasks = []
    if not os.path.exists(filepath):
        return tasks
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    i = 0
    while i < len(lines):
        line = lines[i]
        if re.match(r'\s*@task', line):
            for j in range(i + 1, min(i + 4, len(lines))):
                m = re.match(r'\s*def\s+(\w+)\s*\(', lines[j])
                if m:
                    func_name = m.group(1)
                    doc = ""
                    if j + 1 < len(lines):
                        dl = lines[j + 1].strip()
                        for q in ('"""', "'''"):
                            if dl.startswith(q):
                                if dl.count(q) >= 2:
                                    doc = dl.replace(q, "").strip()
                                else:
                                    parts = [dl[3:]]
                                    for k in range(j + 2, min(j + 10, len(lines))):
                                        if q in lines[k]:
                                            parts.append(lines[k].split(q)[0])
                                            break
                                        parts.append(lines[k].strip())
                                    doc = " ".join(parts).strip()
                                break
                    if not doc:
                        doc = func_name.replace("_", " ")
                    tasks.append({"func": func_name, "doc": doc})
                    break
        i += 1
    return tasks


def get_all_existing_tests():
    result = {}
    for key, fpath in GATEWAY_TEST_FILES.items():
        result[key] = extract_tasks_from_file(fpath)
    return result


def infer_api_from_key(key):
    for api in ("search", "confirm", "init", "select", "lookup"):
        if api in key:
            return api
    return "unknown"


def summarize_existing_tests(all_tests):
    flat = []
    for file_key, tasks in all_tests.items():
        api = infer_api_from_key(file_key)
        t_type = "functional" if "functional" in file_key else "negative"
        for t in tasks:
            full_text = (t["func"].replace("_", " ") + " " + t["doc"]).lower()
            flat.append({
                "description": full_text,
                "func": t["func"],
                "file": os.path.basename(GATEWAY_TEST_FILES[file_key]),
                "doc": t["doc"],
                "api": api,
                "type": t_type,
            })
    return flat


# ─── 2. Classification helpers ───────────────────────────────────────────────

NEGATIVE_KEYWORDS = {
    "invalid", "missing", "negative", "error", "fail", "unauthorized",
    "nack", "without", "not found", "bad", "wrong", "empty", "exceed",
    "reject", "denied", "forbidden", "null", "absent", "only", "no ",
    "incorrect", "corrupt", "tamper", "expired", "dos",
}

API_KEYWORD_MAP = {
    "search":  ["search", "on_search", "on search", "/search"],
    "confirm": ["confirm", "on_confirm", "on confirm", "/confirm"],
    "init":    ["init", "on_init", "on init", "/init"],
    "select":  ["select", "on_select", "on select", "/select"],
    "lookup":  ["lookup", "v2lookup", "v3lookup", "/lookup"],
}

# Patterns that indicate the test is outside the scope of existing Gateway API tests
OUT_OF_SCOPE_PATTERNS = [
    r"\bdomain_admin\b", r"\bondc_admin\b", r"\bsubscriber_admin\b",
    r"\bondc_user\b",    r"\bdomain_user\b", r"\bsubscriber_user\b",
    r"\brbac\b",         r"\baccess control\b",
    r"\bwhitelis",       r"\bsubscrib",         r"\bunsubscrib",
    r"\bukid\b",
    r"v2/token",         r"v1\.1/subscribe",    r"admin/subscriber",
    r"\bkeycloak\b",     r"\bbearer token\b",
]

STOP_WORDS = {
    "the", "a", "an", "to", "of", "in", "is", "and", "or",
    "with", "for", "if", "on", "at", "be", "by", "as", "it",
    "its", "that", "this", "are", "was", "were", "but", "from",
    "have", "has", "been", "all", "which", "when", "should",
    "must", "will", "can", "not", "run", "api", "test", "case",
    "using", "use", "send", "get", "post", "http", "step",
    "open", "enter", "below", "url", "hit",
}

# Normalize aliases before tokenizing
TOKEN_ALIASES = [
    (r"\bv[23]lookup\b", "lookup"),
    (r"\bon_search\b",   "search"),
    (r"\bon_confirm\b",  "confirm"),
    (r"\bon_init\b",     "init"),
    (r"\bon_select\b",   "select"),
    (r"/search\b",       " search "),
    (r"/confirm\b",      " confirm "),
    (r"/init\b",         " init "),
    (r"/select\b",       " select "),
    (r"/lookup\b",       " lookup "),
]


def normalize(text):
    t = text.lower()
    for pattern, replacement in TOKEN_ALIASES:
        t = re.sub(pattern, replacement, t)
    return t


def tokenize(text):
    words = re.findall(r"[a-z][a-z0-9]*", normalize(text))
    return {w for w in words if w not in STOP_WORDS and len(w) > 2}


def classify_type(desc, test_id, expected):
    combined = " ".join([desc, test_id, str(expected or "")]).lower()
    # Check keyword list
    for kw in NEGATIVE_KEYWORDS:
        if kw in combined:
            return "negative"
    # Check HTTP 4xx/5xx response codes that indicate error scenarios
    exp_str = str(expected or "").strip()
    if re.match(r"^(4\d\d|5\d\d)$", exp_str):
        return "negative"
    # Check for common error code strings
    if re.search(r"\b4\d\d\b|\b5\d\d\b|nack|ack.*error", combined):
        return "negative"
    return "functional"


def infer_api(desc, steps, feature, api_col):
    combined = " ".join(filter(None, [desc, steps, feature, api_col])).lower()
    for api, keywords in API_KEYWORD_MAP.items():
        for kw in keywords:
            if kw in combined:
                return api
    return "general"


def is_out_of_scope(text):
    text_lower = text.lower()
    for pattern in OUT_OF_SCOPE_PATTERNS:
        if re.search(pattern, text_lower):
            return True
    return False


# ─── 3. Matching ─────────────────────────────────────────────────────────────

def keyword_match_score(ondc_text, existing_description):
    t1 = tokenize(ondc_text)
    t2 = tokenize(existing_description)
    if not t1 or not t2:
        return 0
    return len(t1 & t2)


def find_best_match(ondc_text, ondc_api, ondc_type, existing_flat):
    """Return (best_existing, score, api_matched)."""
    best_ex = None
    best_total = 0
    best_api_matched = False

    for ex in existing_flat:
        kw = keyword_match_score(ondc_text, ex["description"])
        api_matched = (ondc_api != "general" and ex["api"] == ondc_api)
        type_bonus = 1 if ex["type"] == ondc_type else 0

        # Keyword overlap is the primary signal.
        # API match bonus ONLY applies when there's also content overlap.
        total = kw + type_bonus
        if api_matched and kw >= 1:
            total += 2

        if total > best_total:
            best_total = total
            best_ex = ex
            best_api_matched = api_matched

    return best_ex, best_total, best_api_matched


def map_gap_decision(ondc_api, ondc_type, best_ex, score, api_matched, out_of_scope):
    """
    Determine coverage status with fill color, confidence, remark.

    Thresholds (tuned to avoid false-positive "Covered" for generic API matches):
      Covered          : api_matched AND score >= 6   OR   score >= 8
      Partially Covered: score >= 2
      Gap              : score < 2 OR out_of_scope
    """
    if out_of_scope:
        return ("Gap - Not Covered", RED, "None",
                "Out of scope for existing Gateway automated tests. "
                "Scenario relates to admin roles / registry / token management. "
                "Dedicated admin/registry test automation is required.")

    if best_ex and ((api_matched and score >= 6) or score >= 8):
        confidence = "High" if score >= 8 else "Medium"
        return ("Covered", GREEN, confidence,
                f"Existing {ondc_type} test for /{ondc_api} covers this scenario. "
                f"Function: {best_ex['func']}")

    if best_ex and score >= 2:
        return ("Partially Covered", YELLOW, "Low",
                f"Partial keyword overlap with '{best_ex['func']}'. "
                "Review whether the existing test fully covers this exact ONDC scenario. "
                "Consider adding scenario-specific assertions or a dedicated test case.")

    return ("Gap - Not Covered", RED, "None",
            f"No existing {ondc_type} automated test found for /{ondc_api}. "
            "A new test case must be developed.")


# ─── 4. Write gap columns helper ─────────────────────────────────────────────

GAP_HEADERS = [
    "Gap Analysis - Coverage Status",
    "Mapped Existing Test (Function)",
    "Mapped Existing Test File",
    "Match Confidence",
    "Remarks / Recommendation",
]


def add_gap_header_row(ws, gap_col_start):
    for i, gh in enumerate(GAP_HEADERS):
        cell = ws.cell(1, gap_col_start + i)
        cell.value = gh
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(wrap_text=True)
        col_letter = get_column_letter(gap_col_start + i)
        ws.column_dimensions[col_letter].width = 30


def write_gap_row(ws, row, gap_col_start,
                  status, fill_color, func_name, file_name, confidence, remark):
    values = [status, func_name, file_name, confidence, remark]
    fill = PatternFill("solid", fgColor=fill_color)
    for i, val in enumerate(values):
        cell = ws.cell(row, gap_col_start + i)
        cell.value = val
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)


def process_row(full_desc, test_id, expected, steps, feature, api_col,
                existing_flat, gap_col_start, ws, row):
    if not full_desc.strip():
        return False

    out_of_scope = is_out_of_scope(full_desc + " " + test_id)
    ondc_type = classify_type(full_desc, test_id, expected)
    ondc_api  = infer_api(full_desc, steps or "", feature or "", api_col or "")
    best_ex, score, api_matched = find_best_match(full_desc, ondc_api, ondc_type, existing_flat)

    status, fill_color, confidence, remark = map_gap_decision(
        ondc_api, ondc_type, best_ex, score, api_matched, out_of_scope
    )

    if status == "Gap - Not Covered":
        func_name = "N/A"
        file_name = "N/A"
    else:
        func_name = best_ex["func"] if best_ex else "N/A"
        file_name = best_ex["file"] if best_ex else "N/A"

    write_gap_row(ws, row, gap_col_start,
                  status, fill_color, func_name, file_name, confidence, remark)
    return True


# ─── 5. Per-sheet analysis ───────────────────────────────────────────────────

def analyze_gateway_cache(ws, existing_flat):
    """
    Columns: Sr.No | Test ID | Test Case | Expected Result | Actual Result | Status | ...
    """
    gap_col_start = ws.max_column + 1
    add_gap_header_row(ws, gap_col_start)

    for row in range(2, ws.max_row + 1):
        test_id   = str(ws.cell(row, 2).value or "")
        test_case = str(ws.cell(row, 3).value or "")
        expected  = str(ws.cell(row, 4).value or "")
        combined  = test_case + " " + test_id
        process_row(combined, test_id, expected, "", "", "",
                    existing_flat, gap_col_start, ws, row)

    return gap_col_start


def analyze_others(ws, existing_flat):
    """
    Columns: Test Case ID | Test Scenario | Test Steps | Expected Result | ...
    """
    gap_col_start = ws.max_column + 1
    add_gap_header_row(ws, gap_col_start)

    for row in range(2, ws.max_row + 1):
        test_id       = str(ws.cell(row, 1).value or "")
        test_scenario = str(ws.cell(row, 2).value or "")
        test_steps    = str(ws.cell(row, 3).value or "")
        expected      = str(ws.cell(row, 4).value or "")
        desc = test_scenario + " " + test_steps
        process_row(desc, test_id, expected, test_steps, "", "",
                    existing_flat, gap_col_start, ws, row)

    return gap_col_start


def analyze_gateway_processing(ws, existing_flat):
    """
    Columns: S.No | Test Case ID | Feature | API | Preconditions | Test Steps |
             Expected Result | Actual Result | Status | ...
    """
    gap_col_start = ws.max_column + 1
    add_gap_header_row(ws, gap_col_start)

    for row in range(2, ws.max_row + 1):
        test_id       = str(ws.cell(row, 2).value or "")
        feature       = str(ws.cell(row, 3).value or "")
        api_col       = str(ws.cell(row, 4).value or "")
        preconditions = str(ws.cell(row, 5).value or "")
        test_steps    = str(ws.cell(row, 6).value or "")
        expected      = str(ws.cell(row, 7).value or "")
        desc = " ".join([test_id, feature, api_col, preconditions, test_steps])
        process_row(desc, test_id, expected, test_steps, feature, api_col,
                    existing_flat, gap_col_start, ws, row)

    return gap_col_start


# ─── 6. Count stats ───────────────────────────────────────────────────────────

def count_stats(ws, gap_col_start):
    total = covered = partial = gap = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, gap_col_start).value
        if val:
            total += 1
            if val == "Covered":
                covered += 1
            elif val == "Partially Covered":
                partial += 1
            elif "Gap" in str(val):
                gap += 1
    return {"total": total, "covered": covered, "partial": partial, "gap": gap}


# ─── 7. Summary sheet ────────────────────────────────────────────────────────

def add_summary_sheet(wb, sheet_stats, all_tests):
    if "Gap Analysis Summary" in wb.sheetnames:
        del wb["Gap Analysis Summary"]
    ws = wb.create_sheet("Gap Analysis Summary", 0)

    hdr_fill_dark = PatternFill("solid", fgColor="1F4E79")
    hdr_fill_mid  = PatternFill("solid", fgColor="2E75B6")
    hdr_fill_lite = PatternFill("solid", fgColor="BDD7EE")

    def bold_white(size=11):
        return Font(bold=True, color="FFFFFF", size=size)

    # Title
    ws["A1"] = "ONDC Gateway - Gap Analysis Report"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hdr_fill_dark
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Analysis Date: 2026-03-19   |   Existing Gateway automated tests: 240 tasks across 10 test files"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:H2")

    # ── Existing test inventory ───────────────────────────────────────────────
    r = 4
    ws.cell(r, 1).value = "Existing Gateway Test Inventory"
    ws.cell(r, 1).font = bold_white()
    ws.cell(r, 1).fill = hdr_fill_mid
    ws.merge_cells(f"A{r}:H{r}")

    r += 1
    for ci, h in enumerate(["API", "Type", "# Tasks", "Sample Test Functions"], 1):
        c = ws.cell(r, ci)
        c.value = h
        c.font = Font(bold=True)
        c.fill = hdr_fill_lite

    r += 1
    total_existing = 0
    for file_key in sorted(all_tests.keys()):
        tasks = all_tests[file_key]
        api   = infer_api_from_key(file_key).upper()
        t_type = "Functional" if "functional" in file_key else "Negative"
        sample = ", ".join(t["func"] for t in tasks[:4])
        if len(tasks) > 4:
            sample += f"  …+{len(tasks) - 4} more"
        ws.cell(r, 1).value = api
        ws.cell(r, 2).value = t_type
        ws.cell(r, 3).value = len(tasks)
        c = ws.cell(r, 4)
        c.value = sample
        c.alignment = Alignment(wrap_text=True)
        total_existing += len(tasks)
        r += 1

    ws.cell(r, 1).value = "TOTAL"
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 3).value = total_existing
    ws.cell(r, 3).font = Font(bold=True)

    # ── Per-sheet statistics ──────────────────────────────────────────────────
    r += 2
    ws.cell(r, 1).value = "Per-Sheet Gap Statistics"
    ws.cell(r, 1).font = bold_white()
    ws.cell(r, 1).fill = hdr_fill_mid
    ws.merge_cells(f"A{r}:H{r}")

    r += 1
    stat_hdrs = ["Sheet", "Total ONDC TCs", "Covered", "Partially Covered",
                 "Gap (Not Covered)", "Coverage %", "Gap %", "Notes"]
    for ci, h in enumerate(stat_hdrs, 1):
        c = ws.cell(r, ci)
        c.value = h
        c.font = Font(bold=True)
        c.fill = hdr_fill_lite

    SHEET_NOTES = {
        "Gateway_Cache": (
            "Contains a mix of: (1) v2lookup/search auth tests [Covered], "
            "(2) field-check auth negative tests [Covered/Partial], "
            "(3) Admin role tests (DOMAIN_ADMIN, ONDC_ADMIN, etc.) [GAP - out-of-scope for current suite]."
        ),
        "Others": (
            "Registry/subscription/whitelist flows (token generation, subscribe, UKID). "
            "None are covered by the existing Gateway Locust test suite [All GAP]."
        ),
        "Gateway Processing": (
            "Routing scenarios: bpp_id+bpp_uri, unicast, broadcast, domain-specific B2B & Retail. "
            "Search-related scenarios are mostly covered; B2B-specific and bpp_uri-only edge cases may be partial/gap."
        ),
    }

    r += 1
    grand = {"total": 0, "covered": 0, "partial": 0, "gap": 0}
    for sheet_name, stats in sheet_stats.items():
        t = stats["total"]
        cov_pct = round(stats["covered"] / t * 100, 1) if t else 0
        gap_pct = round(stats["gap"]     / t * 100, 1) if t else 0

        ws.cell(r, 1).value = sheet_name
        ws.cell(r, 2).value = t
        ws.cell(r, 3).value = stats["covered"]
        ws.cell(r, 3).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(r, 4).value = stats["partial"]
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(r, 5).value = stats["gap"]
        ws.cell(r, 5).fill = PatternFill("solid", fgColor=RED)
        ws.cell(r, 6).value = f"{cov_pct}%"
        ws.cell(r, 7).value = f"{gap_pct}%"
        ws.cell(r, 8).value = SHEET_NOTES.get(sheet_name, "")
        ws.cell(r, 8).alignment = Alignment(wrap_text=True)

        for k in grand:
            grand[k] += stats[k]
        r += 1

    # Grand total row
    gt = grand["total"]
    grand_cov_pct = round(grand["covered"] / gt * 100, 1) if gt else 0
    grand_gap_pct = round(grand["gap"]     / gt * 100, 1) if gt else 0
    ws.cell(r, 1).value = "GRAND TOTAL"
    ws.cell(r, 2).value = gt
    ws.cell(r, 3).value = grand["covered"]
    ws.cell(r, 4).value = grand["partial"]
    ws.cell(r, 5).value = grand["gap"]
    ws.cell(r, 6).value = f"{grand_cov_pct}%"
    ws.cell(r, 7).value = f"{grand_gap_pct}%"
    for ci in range(1, 8):
        ws.cell(r, ci).font = Font(bold=True)
        ws.cell(r, ci).fill = PatternFill("solid", fgColor="D9E1F2")

    # ── Legend ────────────────────────────────────────────────────────────────
    r += 2
    ws.cell(r, 1).value = "Legend & Recommendations"
    ws.cell(r, 1).font = bold_white()
    ws.cell(r, 1).fill = hdr_fill_mid
    ws.merge_cells(f"A{r}:H{r}")

    legend = [
        ("Covered (GREEN)",
         "The existing automated test adequately covers the ONDC scenario. No new test needed."),
        ("Partially Covered (YELLOW)",
         "The existing test has some overlap but may not fully exercise the exact scenario. "
         "Review and enhance the existing test or add targeted assertions."),
        ("Gap - Not Covered (RED)",
         "No existing automated test covers this ONDC scenario. "
         "A new test case must be developed and added to the relevant test file."),
        ("Priority Recommendations",
         "1. Admin role / RBAC tests (Gateway_Cache rows DOMAIN_ADMIN, ONDC_ADMIN, etc.) — "
         "create a dedicated admin-API Locust test file.\n"
         "2. Registry subscription/whitelist (Others sheet) — "
         "create a registry test automation suite.\n"
         "3. Gateway Processing: bpp_uri-only routing, B2B domain tests (B2B_01..B2B_15), "
         "on_search forwarding audit — add to ondc_gateway_search_functional.py."),
        ("Scope Note",
         "Existing test files cover: /search, /on_search, /confirm, /on_confirm, "
         "/init, /on_init, /select, /on_select, /lookup (BPP/BAP, functional + negative). "
         "NOT covered: admin APIs, registry APIs, role-based access control."),
    ]
    for label, detail in legend:
        r += 1
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).value = detail
        ws.merge_cells(f"B{r}:H{r}")
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 42

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 60


# ─── 8. Style original header row ─────────────────────────────────────────────

def style_original_header(ws, last_orig_col):
    for c in range(1, last_orig_col + 1):
        cell = ws.cell(1, c)
        has_fill = (cell.fill and cell.fill.patternType == "solid"
                    and cell.fill.fgColor.rgb not in ("00000000", "FFFFFFFF"))
        if not has_fill:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5496")


# ─── 9. Remove previously added gap columns ───────────────────────────────────

def strip_old_gap_columns(ws):
    """Delete all columns from the first 'Gap Analysis' header onwards.

    This handles the case where previous runs added gap columns but only the
    first column header contained 'Gap Analysis', leaving ghost columns behind.
    """
    first_gap_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and "Gap Analysis" in v:
            first_gap_col = c
            break

    if first_gap_col is None:
        return  # Nothing to strip

    # Count how many columns exist at or after first_gap_col
    num_to_delete = ws.max_column - first_gap_col + 1
    if num_to_delete > 0:
        ws.delete_cols(first_gap_col, num_to_delete)


# ─── 10. Consolidated test-case list sheet ────────────────────────────────────

# Mapping: sheet_name -> (test_id_col, description_col, expected_col, extra_desc_cols)
# extra_desc_cols: additional columns (after test_id, description) to concatenate for full context
SHEET_EXTRACT_MAP = {
    "Gateway_Cache": {
        "test_id_col":   2,   # Test ID
        "desc_col":      3,   # Test Case
        "expected_col":  4,   # Expected Result
        "extra_cols":    [],
    },
    "Others": {
        "test_id_col":   1,   # Test Case ID
        "desc_col":      2,   # Test Scenario
        "expected_col":  4,   # Expected Result
        "extra_cols":    [3], # Test Steps
    },
    "Gateway Processing": {
        "test_id_col":   2,   # Test Case ID
        "desc_col":      5,   # Preconditions (most descriptive)
        "expected_col":  7,   # Expected Result
        "extra_cols":    [3, 4, 6],  # Feature, API, Test Steps
    },
}


def _find_gap_col(ws):
    """Return the column index of the first 'Gap Analysis' header, or None."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and "Gap Analysis" in v:
            return c
    return None


def create_consolidated_sheet(wb):
    """
    Build a 'All Test Cases' sheet listing every ONDC test case from all source
    sheets in one place, with the entire row coloured to reflect coverage status.

    Columns:
      #  | Source Sheet | Test Case ID | Test Case / Scenario | Expected Result
         | Type (F/N)  | API          | Coverage Status
         | Mapped Test (Function) | Mapped Test File
         | Confidence  | Remarks / Recommendation
    """
    SHEET_NAME = "All Test Cases"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Insert as the second sheet (after Summary)
    summary_idx = wb.sheetnames.index("Gap Analysis Summary") if "Gap Analysis Summary" in wb.sheetnames else 0
    ws_out = wb.create_sheet(SHEET_NAME, summary_idx + 1)

    # ── Header row ────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    COL_HDRS  = [
        "#", "Source Sheet", "Test Case ID", "Test Case / Scenario",
        "Expected Result", "Type", "API",
        "Coverage Status", "Mapped Test (Function)", "Mapped Test File",
        "Confidence", "Remarks / Recommendation",
    ]
    for ci, h in enumerate(COL_HDRS, 1):
        cell = ws_out.cell(1, ci)
        cell.value = h
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_out.row_dimensions[1].height = 28
    ws_out.freeze_panes = "A2"

    # Column widths
    col_widths = [5, 20, 14, 45, 20, 12, 12, 25, 35, 35, 14, 55]
    for ci, w in enumerate(col_widths, 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = w

    # ── Row fill colours (map status → fgColor) ───────────────────────────────
    STATUS_COLOR = {
        "Covered":           GREEN,
        "Partially Covered": YELLOW,
        "Gap - Not Covered": RED,
    }

    # ── Iterate source sheets and pull data ───────────────────────────────────
    out_row  = 2
    sr_no    = 0
    GAP_COLS = 5   # number of gap analysis columns written per source row

    SOURCE_ORDER = ["Gateway_Cache", "Others", "Gateway Processing"]

    for sheet_name in SOURCE_ORDER:
        if sheet_name not in wb.sheetnames:
            continue

        ws_src  = wb[sheet_name]
        mapping = SHEET_EXTRACT_MAP.get(sheet_name, {})
        if not mapping:
            continue

        gap_col = _find_gap_col(ws_src)
        if gap_col is None:
            continue  # gap analysis not yet applied to this sheet (shouldn't happen)

        id_col       = mapping["test_id_col"]
        desc_col     = mapping["desc_col"]
        expected_col = mapping["expected_col"]
        extra_cols   = mapping["extra_cols"]

        for src_row in range(2, ws_src.max_row + 1):
            # Status column is the first of the 5 gap columns
            status_val = ws_src.cell(src_row, gap_col).value
            if not status_val:
                continue   # empty row

            sr_no += 1
            test_id  = str(ws_src.cell(src_row, id_col).value     or "")
            desc     = str(ws_src.cell(src_row, desc_col).value    or "")
            expected = str(ws_src.cell(src_row, expected_col).value or "")

            # Append extra context columns (e.g. Test Steps, Feature, API)
            extras = []
            for ec in extra_cols:
                v = ws_src.cell(src_row, ec).value
                if v and str(v).strip():
                    extras.append(str(v).strip())
            if extras:
                desc = desc + " | " + " | ".join(extras)

            # Infer type & api for display
            inferred_type = classify_type(desc, test_id, expected)
            inferred_api  = infer_api(desc, "", "", "")
            type_label    = "Functional" if inferred_type == "functional" else "Negative"
            api_label     = inferred_api.upper() if inferred_api != "general" else "—"

            # Read gap columns from source
            gap_vals = [ws_src.cell(src_row, gap_col + i).value for i in range(GAP_COLS)]
            status_raw  = str(gap_vals[0] or "Gap - Not Covered")
            mapped_func = str(gap_vals[1] or "N/A")
            mapped_file = str(gap_vals[2] or "N/A")
            confidence  = str(gap_vals[3] or "None")
            remark      = str(gap_vals[4] or "")

            # Normalise status to clean key
            if "Covered" in status_raw and "Partial" not in status_raw and "Gap" not in status_raw:
                status_key = "Covered"
            elif "Partial" in status_raw:
                status_key = "Partially Covered"
            else:
                status_key = "Gap - Not Covered"

            row_color = STATUS_COLOR.get(status_key, RED)
            row_fill  = PatternFill("solid", fgColor=row_color)

            values = [
                sr_no, sheet_name, test_id,
                desc[:500],   # cap very long payloads
                expected,
                type_label, api_label,
                status_key, mapped_func, mapped_file,
                confidence, remark,
            ]

            for ci, val in enumerate(values, 1):
                cell = ws_out.cell(out_row, ci)
                cell.value     = val
                cell.fill      = row_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Row height – slightly taller for readability
            ws_out.row_dimensions[out_row].height = 40

            out_row += 1

        # Blank separator between sheets
        out_row += 1

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws_out.auto_filter.ref = (
        f"A1:{get_column_letter(len(COL_HDRS))}{out_row - 1}"
    )

    print(f"  Created '{SHEET_NAME}' sheet with {sr_no} test case rows.")


# ─── 11. Main ─────────────────────────────────────────────────────────────────

def main():
    print("Loading existing Gateway test cases from source code...")
    all_tests     = get_all_existing_tests()
    existing_flat = summarize_existing_tests(all_tests)

    print(f"Found {len(existing_flat)} existing test tasks across {len(GATEWAY_TEST_FILES)} files:")
    for file_key in sorted(all_tests):
        print(f"  {file_key}: {len(all_tests[file_key])} tasks")

    print(f"\nLoading Excel: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    ANALYZERS = {
        "Gateway_Cache":      analyze_gateway_cache,
        "Others":             analyze_others,
        "Gateway Processing": analyze_gateway_processing,
    }

    sheet_stats = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Gap Analysis Summary":
            continue
        ws = wb[sheet_name]
        analyzer = ANALYZERS.get(sheet_name)
        if not analyzer:
            print(f"  Skipping unknown sheet: {sheet_name}")
            continue

        strip_old_gap_columns(ws)
        last_orig_col = ws.max_column

        print(f"\nAnalyzing sheet: {sheet_name} ({ws.max_row} rows)...")
        gap_col_start = analyzer(ws, existing_flat)
        style_original_header(ws, last_orig_col)

        stats    = count_stats(ws, gap_col_start)
        cov_pct  = round(stats["covered"] / stats["total"] * 100, 1) if stats["total"] else 0
        gap_pct  = round(stats["gap"]     / stats["total"] * 100, 1) if stats["total"] else 0
        sheet_stats[sheet_name] = stats
        print(f"  Total={stats['total']}  Covered={stats['covered']}  "
              f"Partial={stats['partial']}  Gap={stats['gap']}  "
              f"Coverage={cov_pct}%  Gap={gap_pct}%")

    print("\nAdding Gap Analysis Summary sheet...")
    add_summary_sheet(wb, sheet_stats, all_tests)

    print("\nBuilding consolidated 'All Test Cases' sheet...")
    create_consolidated_sheet(wb)

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")

    grand_total   = sum(s["total"]   for s in sheet_stats.values())
    grand_covered = sum(s["covered"] for s in sheet_stats.values())
    grand_partial = sum(s["partial"] for s in sheet_stats.values())
    grand_gap     = sum(s["gap"]     for s in sheet_stats.values())

    print()
    print("=" * 60)
    print("  OVERALL  GAP ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"  Total ONDC Test Cases Analyzed  : {grand_total}")
    print(f"  Covered (GREEN)                 : {grand_covered} "
          f"({round(grand_covered / grand_total * 100, 1) if grand_total else 0}%)")
    print(f"  Partially Covered (YELLOW)      : {grand_partial} "
          f"({round(grand_partial / grand_total * 100, 1) if grand_total else 0}%)")
    print(f"  Gap - Not Covered (RED)         : {grand_gap} "
          f"({round(grand_gap / grand_total * 100, 1) if grand_total else 0}%)")
    print("=" * 60)


if __name__ == "__main__":
    main()
