"""html_to_executive_csv.py

Reads ONDC API test-result HTML files from a folder and produces an
executive-summary Excel workbook.

For each HTML file the script:
  1. Finds every individual test card (pass / fail)
  2. Reads the URL inside each card to identify the real API path  (e.g. /confirm)
  3. Groups cards by API path and counts totals, passes, fails, and bug severity
  4. Aggregates counts across all HTML files by API path
  5. Writes one Excel row per unique API path

Bug severity (applied only to FAIL cards):
  High   – HTTP 401 / 403 / 5xx  →  auth failure or server crash
  Low    – HTTP 400 / 404 / 409 / 422  →  bad-request / not-found
  Medium – HTTP 2xx on a FAIL card  →  functional regression (got success when error expected)
           or any unrecognised code

Usage
-----
  python func_test_scripts/html_to_executive_csv.py                  # reads results/api_reports/
  python func_test_scripts/html_to_executive_csv.py --input folder/
  python func_test_scripts/html_to_executive_csv.py --input folder/ --output report.xlsx
"""
from __future__ import annotations

import argparse
import hashlib
import html as html_lib
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Regexes ───────────────────────────────────────────────────────────────────

# Opening tag of a test card
_RE_CARD_OPEN = re.compile(r'<div\s+class="card\s+(pass|fail)"', re.IGNORECASE)

# Validates this is a supported report (has summary score-cards)
_RE_SCARD = re.compile(r'class="scard\s+(total|passed|failed)', re.IGNORECASE)

# HTTP status chip inside a card header:  <span class="chip sc-2xx">HTTP 200</span>
_RE_CHIP = re.compile(r'<span\s+class="chip\s+(sc-[^"\s]+)[^"]*"\s*>([^<]+)</span>')

# data-suite attribute on the card opening div — used to detect test type
_RE_CARD_SUITE = re.compile(r'data-suite="([^"]*)"', re.IGNORECASE)
# Security-related keywords in a test name — used to promote negative-test bypass to High
_RE_SECURITY_TC = re.compile(
    r'(?:dos|denial.of.service|large.payload|flood|oversize|'
    r'expired.sig|invalid.sig|replay|'
    r'expired.timestamp|invalid.timestamp|'
    r'auth.fail|auth.missing|missing.auth|invalid.auth|auth.bypass|'
    r'sql.inject|xss|inject|traversal|fuzz)',
    re.IGNORECASE,
)
# URL after "URL:" or "Endpoint:" label:
#   <strong>URL:</strong>
#       http://host/path
# The \s+ handles the newline + indentation between </strong> and the URL.
_RE_URL_FIELD = re.compile(
    r'(?:URL|Endpoint):</strong[^>]*>\s+(https?://[^\s<"&]+)',
    re.IGNORECASE,
)

# <title> tag
_RE_TITLE = re.compile(r'<title[^>]*>(.*?)</title>', re.IGNORECASE | re.DOTALL)

# Test case identifier embedded in card text:  TC005_Confirm_Missing_BPP_ID
_RE_TC_NAME = re.compile(r'\b(TC[\w.-]+)', re.IGNORECASE)


# ── Card slicing ──────────────────────────────────────────────────────────────

def _slice_cards(html_content: str) -> list[tuple[str, bool, str]]:
    """
    Return a list of (status, is_negative, card_html) for every test card.
    - status      : 'pass' or 'fail'
    - is_negative : True if the test is a negative/validation test
    - card_html   : full HTML chunk for this card

    Test type is detected from the card's data-suite attribute:
      data-suite contains 'negative'  →  negative test
      anything else (functional, performance, …)  →  positive/functional test
    Falls back to the test name (TC-N- prefix or 'Negative' keyword).
    """
    positions = [
        (m.start(), m.group(1).lower())
        for m in _RE_CARD_OPEN.finditer(html_content)
    ]
    result = []
    for i, (start, status) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(html_content)
        chunk = html_content[start:end]

        # Detect test type from data-suite or test-name signals
        suite_m = _RE_CARD_SUITE.search(chunk[:400])   # only scan opening tag area
        if suite_m:
            is_negative = 'negative' in suite_m.group(1).lower()
        else:
            lower = chunk[:400].lower()
            is_negative = bool(
                re.search(r'tc[-_]n[-_]', lower) or
                '| negative' in lower or
                '· negative' in lower or
                'negative –' in lower or
                'negative -' in lower
            )

        result.append((status, is_negative, chunk))
    return result


# ── Per-card helpers ──────────────────────────────────────────────────────────

def _card_api_path(card_html: str, fallback: str) -> str:
    """Extract the API path from the card's URL/Endpoint field, or return fallback."""
    m = _RE_URL_FIELD.search(card_html)
    if m:
        path = urlparse(m.group(1).strip()).path.rstrip('/')
        if path and path != '/':
            return path
    return fallback


def _classify_severity(card_html: str, is_negative: bool) -> str:
    """
    Classify a FAIL card using both HTTP status and test type.

    High   – Server-side error (5xx) on a valid request
           – Auth failure (401/403) on a functional test
           – Security validation bypass: negative test with security keywords
             (expired/invalid signature, expired timestamp, DoS payload, auth
             missing/failure, injection) accepted with 2xx → critical gap
    Medium – Negative test accepted invalid input (2xx) — non-security
           – Functional test got 4xx (non-auth) : API rejected a valid request
           – Server-side error (5xx) on a negative test
           – Auth failure (401/403) on a negative test (ambiguous)
    Low    – Rate-limit / transient infra (429, 503, 504)
    """
    m = _RE_CHIP.search(card_html)
    chip_class = m.group(1) if m else ''
    chip_text  = m.group(2) if m else ''

    code_m = re.search(r'\d{3}', chip_text)
    code   = int(code_m.group()) if code_m else 0

    # ── Environment / infrastructure issues ─────────────────────────────────
    if code in (429, 503, 504):
        return 'low'

    # ── Server-side errors ────────────────────────────────────────────────────────
    if 'sc-5xx' in chip_class or (500 <= code <= 599):
        # Server-side error on a valid request → high; on invalid input → medium
        return 'high' if not is_negative else 'medium'

    # ── Auth failures ────────────────────────────────────────────────────────
    if code in (401, 403):
        # Auth failure blocking a valid functional request → high
        # Auth failure on a negative test (expected) shouldn't be a FAIL normally,
        # but if it is, treat as medium.
        return 'high' if not is_negative else 'medium'

    # ── Wrong behaviour ──────────────────────────────────────────────────────
    if is_negative and ('sc-2xx' in chip_class or (200 <= code <= 299)):
        # Security-related negative test bypassed → critical, promote to High
        if _RE_SECURITY_TC.search(card_html[:800]):
            return 'high'
        # Non-security validation gap → Medium
        return 'medium'

    if not is_negative and 'sc-4xx' in chip_class:
        # Functional test rejected for wrong reason (bad request, not found, …)
        return 'medium'

    return 'medium'   # fallback


# ── Bug-detail extraction helpers ─────────────────────────────────────────

# Matches:  <div class="col-label">LABEL</div>\n<pre class="json-block">CONTENT</pre>
_RE_COL_PRE = re.compile(
    r'<div\s+class="col-label">\s*([^<]+?)\s*</div>\s*'
    r'<pre[^>]*>([\s\S]*?)</pre>',
    re.IGNORECASE,
)

# TC name from dedicated span
_RE_TC_SPAN = re.compile(r'<span\s+class="tc-name">([^<]+)</span>', re.IGNORECASE)

# Splits card into Request section and Response section
_RE_REQ_SECTION = re.compile(
    r'class="section-title req-title".*?(?=class="section-title res-title"|\Z)',
    re.IGNORECASE | re.DOTALL,
)
_RE_RES_SECTION = re.compile(
    r'class="section-title res-title".*',
    re.IGNORECASE | re.DOTALL,
)


def _clean_html(raw: str, max_len: int = 1200) -> str:
    """Strip tags, unescape HTML entities, normalise whitespace."""
    text = re.sub(r'<[^>]+>', ' ', raw)
    text = html_lib.unescape(text)
    return re.sub(r'\s+', ' ', text).strip()[:max_len]


def _col_label_content(html_section: str, *labels: str, max_len: int = 1200) -> str:
    """Return unescaped text of the first col-label <pre> matching any of the given labels."""
    for m in _RE_COL_PRE.finditer(html_section):
        lbl = m.group(1).strip().lower()
        for label in labels:
            if label.lower() in lbl:
                return html_lib.unescape(m.group(2)).strip()[:max_len]
    return ''


def _bug_record(card_html: str, api: str, severity: str, source: str) -> dict:
    """Build a bug-detail dict from a single FAIL card."""
    tc_m   = _RE_TC_SPAN.search(card_html)
    chip_m = _RE_CHIP.search(card_html)

    tc_name = tc_m.group(1).strip() if tc_m else (_RE_TC_NAME.search(card_html[:600]).group(1) if _RE_TC_NAME.search(card_html[:600]) else '')
    result  = chip_m.group(2).strip() if chip_m else 'Unknown'

    req_m  = _RE_REQ_SECTION.search(card_html)
    res_m  = _RE_RES_SECTION.search(card_html)
    req_section = req_m.group(0) if req_m else card_html
    res_section = res_m.group(0) if res_m else ''

    body     = _col_label_content(req_section,  'body (json)', 'body')
    headers  = _col_label_content(req_section,  'headers')
    response = _col_label_content(res_section,  'body (json)', 'body') or result

    return {
        'classification': severity.capitalize(),
        'api':            api,
        'tc_name':        tc_name,
        'headers':        headers,
        'body':           body,
        'result':         response,
        'http_code':      result,   # e.g. "HTTP 200"
        'source':         source,
    }


# ── Report-level fallback API name ────────────────────────────────────────────

def _report_level_api(html_content: str) -> str:
    """
    Extract an API identifier from the hero/header section of the report
    (used as fallback when individual cards have no URL field).
    Tries the Endpoint/URL field first, then the <title>.
    """
    first_card = _RE_CARD_OPEN.search(html_content)
    hero = html_content[: first_card.start()] if first_card else html_content

    m = _RE_URL_FIELD.search(hero)
    if m:
        path = urlparse(m.group(1).strip()).path.rstrip('/')
        if path and path != '/':
            return path

    m = _RE_TITLE.search(hero)
    if m:
        title = re.sub(r'&[a-zA-Z#\d]+;', ' ', m.group(1))
        title = re.sub(r'\s*[-\u2013\u2014]?\s*API Test Report.*', '', title, flags=re.IGNORECASE)
        title = re.sub(r'^ONDC\s+', '', title, flags=re.IGNORECASE)
        cleaned = title.strip()
        if cleaned:
            return cleaned

    return 'Unknown'


# ── Main parse function ───────────────────────────────────────────────────────

def parse_report(path: Path) -> list[dict]:
    """
    Parse one HTML report file.

    Returns a list of dicts — one per unique API path found in the file:
      { api, total, passed, failed, high, medium, low }

    Returns an empty list if the file is not a recognised format.
    """
    try:
        html_content = path.read_text(encoding='utf-8', errors='replace')
    except OSError as exc:
        print(f"  WARNING: cannot read {path.name}: {exc}")
        return []

    if not _RE_SCARD.search(html_content):
        print(f"  Skipping {path.name} — not a recognised HTML test report.")
        return []

    fallback = _report_level_api(html_content)
    cards    = _slice_cards(html_content)

    if not cards:
        # No card elements found — shouldn't happen but gracefully fall back
        # to the summary-card totals with the report-level API name.
        def _int(pat: re.Pattern) -> int:
            mm = pat.search(html_content)
            return int(mm.group(1)) if mm else 0
        total = re.search(r'class="scard\s+total[^"]*"[^>]*>\s*<div class="val">(\d+)', html_content, re.I)
        passed = re.search(r'class="scard\s+passed[^"]*"[^>]*>\s*<div class="val">(\d+)', html_content, re.I)
        failed = re.search(r'class="scard\s+failed[^"]*"[^>]*>\s*<div class="val">(\d+)', html_content, re.I)
        return [{
            'api':    fallback,
            'source': path.name,
            'total':  int(total.group(1))  if total  else 0,
            'passed': int(passed.group(1)) if passed else 0,
            'failed': int(failed.group(1)) if failed else 0,
            'high': 0, 'medium': 0, 'low': 0,
        }], []

    # Group by API path
    groups: dict[str, dict] = defaultdict(
        lambda: {'total': 0, 'passed': 0, 'failed': 0, 'high': 0, 'medium': 0, 'low': 0}
    )
    bugs: list[dict] = []
    for status, is_negative, card_html in cards:
        api = _card_api_path(card_html, fallback)
        g   = groups[api]
        g['total'] += 1
        if status == 'pass':
            g['passed'] += 1
        else:
            severity = _classify_severity(card_html, is_negative)
            g['failed'] += 1
            g[severity] += 1
            bugs.append(_bug_record(card_html, api, severity, path.name))

    # If the fallback name (from title) is still present as a group key, it means
    # some cards had no URL field and were bucketed under a human-readable label
    # instead of an API path.  When real URL-path groups exist, drop the fallback
    # group entirely — it is not a real API and its counts cannot be attributed.
    if fallback in groups and not fallback.startswith('/'):
        url_paths = [k for k in groups if k.startswith('/')]
        if len(url_paths) == 1:
            # Absorb into the single real path
            target = groups[url_paths[0]]
            fb = groups.pop(fallback)
            for key in ('total', 'passed', 'failed', 'high', 'medium', 'low'):
                target[key] += fb[key]
        elif len(url_paths) > 1:
            # Multiple real paths — cannot attribute, just discard the fallback row
            groups.pop(fallback)

    return [{'api': api, 'source': path.name, **data} for api, data in sorted(groups.items())], bugs


# ── Aggregation ───────────────────────────────────────────────────────────────

def aggregate(rows: list[dict]) -> list[dict]:
    """Merge rows with the same API name, summing all numeric counts."""
    agg: dict[str, dict] = defaultdict(
        lambda: {'total': 0, 'passed': 0, 'failed': 0, 'high': 0, 'medium': 0, 'low': 0, 'sources': []}
    )
    for row in rows:
        for key in ('total', 'passed', 'failed', 'high', 'medium', 'low'):
            agg[row['api']][key] += row[key]
        src = row.get('source', '')
        if src and src not in agg[row['api']]['sources']:
            agg[row['api']]['sources'].append(src)
    result = []
    for api, data in sorted(agg.items(), key=lambda x: x[0].lower()):
        sources = data.pop('sources')
        result.append({'api': api, 'source': ', '.join(sources), **data})
    return result


# ── Excel writer ──────────────────────────────────────────────────────────────

_NUM_COLS   = 8
_COL_WIDTHS = [32, 24, 12, 12, 14, 16, 14, 40]
_DARK_BLUE  = '1F3864'
_LT_GREEN   = 'E2EFDA'
_LT_RED     = 'FFDCE2'
_RED_CLR    = 'C00000'

_BUG_HEADERS = ['#', 'Classification', 'API Tested', 'Test Case Name',
                'Headers Used', 'Request Body', 'Result Returned', 'Returned Value', 'Source File']
_BUG_WIDTHS  = [5, 16, 22, 32, 48, 55, 30, 16, 38]


def _border() -> Border:
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def write_bug_sheet(wb: openpyxl.Workbook, bug_records: list[dict]) -> None:
    """Add a 'Bug Details' worksheet listing every individual fail card."""
    ws = wb.create_sheet('Bug Details')

    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill(start_color=_DARK_BLUE, end_color=_DARK_BLUE, fill_type='solid')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_top  = Alignment(horizontal='left', vertical='top', wrap_text=True)
    bdr       = _border()

    ws.row_dimensions[1].height = 26
    for col, text in enumerate(_BUG_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr

    _SEV_STYLE = {
        'High':   ('FF0000', 'FFF0F0'),
        'Medium': ('FF8C00', 'FFF8EC'),
        'Low':    ('888800', 'FFFFF0'),
    }

    for ri, bug in enumerate(bug_records, 2):
        sev = bug['classification']
        font_clr, bg_clr = _SEV_STYLE.get(sev, ('000000', 'FFFFFF'))
        row_fill = PatternFill(start_color=bg_clr, end_color=bg_clr, fill_type='solid')
        sev_font = Font(bold=True, color=font_clr, size=10)
        ws.row_dimensions[ri].height = 75

        vals = [ri - 1, sev, bug['api'], bug['tc_name'],
                bug['headers'], bug['body'], bug['result'], bug['http_code'], bug['source']]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border    = bdr
            c.fill      = row_fill
            c.alignment = center if col in (1, 2) else left_top
            if col == 2:
                c.font = sev_font
            if col == 9 and val:   # Source File — add hyperlink
                c.hyperlink = str(val)
                c.font = Font(color='0563C1', underline='single', size=10)

    for col, width in enumerate(_BUG_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def write_excel(rows: list[dict], bug_records: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary Report'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color=_DARK_BLUE, end_color=_DARK_BLUE, fill_type='solid')
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left',   vertical='center')
    bdr      = _border()

    # ── Two-row header ────────────────────────────────────────────────────────
    # Row 1: API Tested | Total | Pass | Fail | ←── Bugs ───→ | Source File
    # Row 2:  (merged)  |  (m)  | (m)  | (m)  | High Med Low  |   (merged)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    # Columns that span both header rows vertically
    for col, text in [(1, 'API Tested'), (2, 'Total # of Test Cases'),
                      (3, 'Pass'), (4, 'Fail'), (8, 'Source File')]:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        c = ws.cell(row=1, column=col, value=text)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
        ws.cell(row=2, column=col).border = bdr
        ws.cell(row=2, column=col).fill   = hdr_fill

    # "Bugs" group header spanning cols 5-7
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    bugs_c = ws.cell(row=1, column=5, value='Bugs')
    bugs_c.font = hdr_font; bugs_c.fill = hdr_fill; bugs_c.alignment = center; bugs_c.border = bdr
    for col in (6, 7):
        ws.cell(row=1, column=col).fill = hdr_fill
        ws.cell(row=1, column=col).border = bdr

    # Sub-headers: High / Medium / Low
    for col, text in [(5, 'High'), (6, 'Medium'), (7, 'Low')]:
        c = ws.cell(row=2, column=col, value=text)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr

    for ri, row in enumerate(rows, 3):
        vals = [row['api'], row['total'], row['passed'], row['failed'],
                row['high'], row['medium'], row['low'], row.get('source', '')]
        ws.row_dimensions[ri].height = 22

        if row['failed'] == 0 and row['total'] > 0:
            fill = PatternFill(start_color=_LT_GREEN, end_color=_LT_GREEN, fill_type='solid')
        elif row['failed'] > 0:
            fill = PatternFill(start_color=_LT_RED,   end_color=_LT_RED,   fill_type='solid')
        else:
            fill = None

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border    = bdr
            c.alignment = left if col == 1 else center
            if fill:
                c.fill = fill
            if col in (4, 5, 6, 7) and isinstance(val, int) and val > 0:
                c.font = Font(bold=(col == 4), color=_RED_CLR)
            if col == 8:   # Source File — left-aligned, smaller font
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                # Add hyperlink when there is exactly one source file
                sources = [s.strip() for s in str(val).split(',')] if val else []
                if len(sources) == 1 and sources[0]:
                    c.hyperlink = sources[0]
                    c.font = Font(color='0563C1', underline='single')

    for col, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A3'

    # ── Severity legend ───────────────────────────────────────────────────────
    legend_start = len(rows) + 4          # 2 header rows + data + 1 blank gap

    legend_title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
    legend_title_font = Font(bold=True, color='FFFFFF', size=10)

    title_cell = ws.cell(row=legend_start, column=1, value='Bug Severity Classification')
    title_cell.font      = legend_title_font
    title_cell.fill      = legend_title_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[legend_start].height = 22
    # Extend fill across all columns
    for col in range(2, _NUM_COLS + 1):
        c = ws.cell(row=legend_start, column=col)
        c.fill = legend_title_fill

    legend_rows = [
        ('High',   'FF0000', 'FFF0F0',
         'Server-side error (5xx) on a valid request  |  Auth failure (401/403) blocking a valid call  |  '
         'Security validation bypass: negative test with expired/invalid signature, expired timestamp, '
         'DoS payload, or auth failure accepted with HTTP 200'),
        ('Medium', 'FF8C00', 'FFF8EC',
         'Negative test accepted invalid input (2xx when error expected)  |  '
         'Valid request rejected (4xx)  |  Server-side error (5xx) on invalid input  |  '
         'Auth failure (401/403) on a negative test (ambiguous outcome)  |  '
         'Unclassified failure (no HTTP status detected)'),
        ('Low',    '888800', 'FFFFF0',
         'Rate-limit / environment issue (429 / 503 / 504) — transient infrastructure issue, not an application defect'),
    ]

    for offset, (label, font_clr, bg_clr, description) in enumerate(legend_rows, 1):
        row_num = legend_start + offset
        # Taller row for Medium (long description), normal for others
        ws.row_dimensions[row_num].height = 48 if label == 'Medium' else 22

        sev_fill = PatternFill(start_color=bg_clr, end_color=bg_clr, fill_type='solid')
        sev_font = Font(bold=True, color=font_clr, size=10)

        lbl_cell = ws.cell(row=row_num, column=1, value=f'  {label}')
        lbl_cell.font      = sev_font
        lbl_cell.fill      = sev_fill
        lbl_cell.alignment = Alignment(horizontal='left', vertical='center')

        desc_cell = ws.cell(row=row_num, column=2, value=description)
        desc_cell.font      = Font(size=9, color='444444')
        desc_cell.fill      = sev_fill
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Merge description across remaining columns
        ws.merge_cells(
            start_row=row_num, start_column=2,
            end_row=row_num,   end_column=_NUM_COLS
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_bug_sheet(wb, bug_records)
    wb.save(output_path)


# ── CLI ───────────────────────────────────────────────────────────────────────


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description='Executive summary Excel from ONDC API test HTML files.'
    )
    p.add_argument('--input', '-i', required=True,
                   help='Folder containing HTML test report files')
    p.add_argument('--output', '-o', default=None,
                   help='Output .xlsx path (default: <folder>/ONDC-QA-Summary-Report-<date>.xlsx)')
    return p.parse_args()


def main() -> None:
    args   = parse_args()
    folder = Path(args.input)

    if not folder.is_dir():
        print(f"ERROR: folder not found: {folder}", file=sys.stderr)
        sys.exit(1)

    html_files = sorted(f for f in folder.iterdir()
                        if f.is_file() and f.suffix.lower() == '.html')
    if not html_files:
        print(f"No HTML files found in: {folder}", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(html_files)} HTML file(s) in '{folder}':")
    raw_rows: list[dict] = []
    all_bugs: list[dict] = []
    seen_hashes: dict[str, str] = {}   # hash → first filename
    for f in html_files:
        content_hash = hashlib.md5(f.read_bytes()).hexdigest()
        if content_hash in seen_hashes:
            print(f"  SKIPPED (duplicate of '{seen_hashes[content_hash]}'): {f.name}")
            continue
        seen_hashes[content_hash] = f.name
        print(f"  Parsing: {f.name}")
        summary, bugs = parse_report(f)
        raw_rows.extend(summary)
        all_bugs.extend(bugs)

    if not raw_rows:
        print("No reports could be parsed.", file=sys.stderr)
        sys.exit(1)

    # Aggregate all test cases by API path across all files
    rows = aggregate(raw_rows)

    out = Path(args.output) if args.output else folder / f'ONDC-QA-Summary-Report-{date.today().strftime("%Y%m%d")}.xlsx'

    # Sort bugs: High first, then Medium, then Low; within each group by API name
    _SEV_ORDER = {'High': 0, 'Medium': 1, 'Low': 2}
    all_bugs.sort(key=lambda b: (_SEV_ORDER.get(b['classification'], 9), b['api'].lower()))

    write_excel(rows, all_bugs, out)

    print(f"\nExecutive summary saved: {out}")
    print(f"Unique APIs: {len(rows)}\n")
    print(f"  {'API Tested':<35} {'Total':>7} {'Pass':>6} {'Fail':>6} {'High':>6} {'Med':>6} {'Low':>6}")
    print(f"  {'-'*35} {'-'*7} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}")
    for r in rows:
        print(f"  {r['api']:<35} {r['total']:>7} {r['passed']:>6} "
              f"{r['failed']:>6} {r['high']:>6} {r['medium']:>6} {r['low']:>6}")


if __name__ == '__main__':
    main()



if __name__ == '__main__':
    main()
